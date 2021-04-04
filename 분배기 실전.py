import discord
import asyncio
import random
import openpyxl
import datetime

client = discord.Client()

token = "ODI3NzE0ODI4MDE3MDc0MjM2.YGfDgQ.lE3hWlGYnGb1SCLNoCFv4CxLkgQ"

@client.event
async def on_ready():

    print(client.user.name)
    print('코드분배기 준비 완료')
    game = discord.Game('월급 가챠')
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):
    if message.content.startswith("!유저등록"):
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active
        user = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = user[1]
                sheet["B" + str(i)].value = user[2]
                break
        await message.channel.send("유저정보가 성공적으로 등록되었습니다")
        file.save("유저정보.xlsx")
        
    
    if message.content.startswith("!코드등록"):
         file = openpyxl.load_workbook("유저정보.xlsx")
         sheet = file.active
         code = message.content.split(" ")
         codes = code[1].split("/")
         random.shuffle(codes)
         for i in range(0, len(codes)):
             await message.channel.send(codes[i])
         for i in range(1, len(codes)):
            if sheet["C" + str(i)].value == "?":
                sheet["C" + str(i)].value = codes[i]
            elif sheet["C" + str(i)].value != "?":
                break

            file.save("유저정보.xlsx")


    if message.content.startswith("!파일 초기화"):
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active
        for i in range(1, 51):
            if sheet["A" + str(i)].value != "-":
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = " "
                sheet["C" + str(i)].value = "?"
                sheet["D" + str(i)].value = " "
            elif sheet["A" + str(i)].value == "-":
                await message.channel.send("파일 초기화 완료")
                break
        file.save("유저정보.xlsx")
   
    if message.content.startswith("!월급가챠"):
        file = openpyxl.load_workbook("유저정보.xlsx")
        sheet = file.active
        na = message.content.split(" ")
        user = message.author
        await message.channel.send(user)
        for i in range(1, 51):
            code = sheet["C" + str(i)].value
            if sheet["B" + str(i)].value == na[1]:
                sheet["D" + str(i)].value = "수령완료"
                await message.author.send("네이비썰 촬영에 참여해주셔서 감사합니다")
                await message.author.send(code)
                break
            elif sheet["D" + str(i)].value == "수령완료":
                await message.channel.send("이미 수령하였습니다")
                break
        file.save("유저정보.xlsx")


client.run(token)